"""Reads a Sample Buy workbook back into the same row shape
group_avails_into_rows() produces, so a buyer's manual edits -- changed day
quantities, adjusted rates, added or deleted rows -- flow straight into the
Strata export without needing the original rate card XML files again.

The reader is driven entirely by the header row's text labels rather than
hardcoded column positions, so it keeps working if excel_export.py's column
order ever changes. A row counts as a buy line if its Program cell is
non-empty -- subtotal and "MARKET TOTAL" rows never have one, so they're
skipped automatically with no special-casing, and a buyer deleting a row
entirely just means it's absent here too.
"""

from openpyxl import load_workbook

DAY_HEADER_TO_NAME = {"M": "Mon", "T": "Tue", "W": "Wed", "Th": "Thu", "F": "Fri", "Sa": "Sat", "Su": "Sun"}
REQUIRED_HEADERS = ["Daypart", "Station", "Program", "Time", "Length", "Rate", "Rating"]


def _clock_to_min(label):
    """Inverse of excel_export._min_to_clock: '7:00a' / '11:30p' -> minutes since midnight."""
    label = label.strip().lower()
    hour_text, rest = label.split(":")
    minute_text, suffix = rest[:-1], rest[-1]
    hour, minute = int(hour_text), int(minute_text)
    if suffix == "a":
        if hour == 12:
            hour = 0
    elif suffix == "p":
        if hour != 12:
            hour += 12
    else:
        raise ValueError(f"expected a trailing 'a' or 'p', got {label!r}")
    return hour * 60 + minute


def _find_header_row(ws):
    """Scans for the grid header row by its known label set, rather than
    assuming a fixed row number -- the station summary block above it grows
    or shrinks with the number of stations in the buy."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = {cell.value: cell.column for cell in row if cell.value is not None}
        if {"Daypart", "Station", "Program"} <= values.keys():
            return row[0].row, values
    raise ValueError("Could not find the Sample Buy grid header row (looking for Daypart/Station/Program columns)")


def read_sample_buy_rows(path, sheet_name="Sample Buy"):
    """Returns a list of row dicts (station, program, start_min, end_min,
    day_counts, rate, rating, spot_length, daypart_name) for every buy line
    still present in the sheet -- reflecting whatever a buyer edited by
    hand. Raises ValueError if the file doesn't look like a Sample Buy
    export."""
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No {sheet_name!r} sheet found -- is this a Sample Buy workbook?")
    ws = wb[sheet_name]

    header_row, col = _find_header_row(ws)
    missing = [h for h in REQUIRED_HEADERS if h not in col]
    if missing:
        raise ValueError(f"Sample Buy sheet is missing expected column(s): {', '.join(missing)}")
    day_cols = {name: col[label] for label, name in DAY_HEADER_TO_NAME.items() if label in col}

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        program = ws.cell(row=r, column=col["Program"]).value
        if not program:
            continue  # subtotal / MARKET TOTAL / blank separator rows all lack a Program

        time_label = ws.cell(row=r, column=col["Time"]).value or ""
        try:
            start_label, end_label = time_label.split("-", 1)
            start_min = _clock_to_min(start_label)
            end_min = _clock_to_min(end_label)
        except (ValueError, AttributeError) as e:
            raise ValueError(f"Row {r}: couldn't parse time range {time_label!r} (expected e.g. '7:00a-8:00a'): {e}")

        day_counts = {}
        for day_name, day_col in day_cols.items():
            value = ws.cell(row=r, column=day_col).value
            day_counts[day_name] = int(value) if value else 0

        rows.append(
            {
                "daypart_name": ws.cell(row=r, column=col["Daypart"]).value or "",
                "station": ws.cell(row=r, column=col["Station"]).value,
                "program": program,
                "start_min": start_min,
                "end_min": end_min,
                "day_counts": day_counts,
                "rate": ws.cell(row=r, column=col["Rate"]).value or 0,
                "rating": ws.cell(row=r, column=col["Rating"]).value or 0,
                "spot_length": ws.cell(row=r, column=col["Length"]).value or "00:00:30",
            }
        )
    return rows
