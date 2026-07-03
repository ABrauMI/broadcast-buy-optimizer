from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .classify import daypart_code

# ---- GPS Impact brand kit (Brand Guidelines 2026) ----
NAVY = "323B51"
PALE_BLUE = "BED7D5"
BLUE = "3D6A91"
RED = "DE5E4E"
WHITE = "FFFFFF"
FONT_BODY = "Figtree"
FONT_DISPLAY = "Superior Title Bold"
LOGO_PATH = Path(__file__).parent / "assets" / "gps_impact_logo_white.png"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name=FONT_BODY, color=WHITE, bold=True)
SECTION_FONT = Font(name=FONT_BODY, bold=True, size=12, color=NAVY)
TITLE_FONT = Font(name=FONT_DISPLAY, bold=True, size=16, color=WHITE)
BOLD_FONT = Font(name=FONT_BODY, bold=True)
DAY_BOUGHT_FONT = Font(name=FONT_BODY, color=WHITE, bold=True)
DAY_UNBOUGHT_FONT = Font(name=FONT_BODY, color="595959")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

CATEGORY_ORDER = [
    "Early News",
    "Noon News",
    "Evening News",
    "Late News",
    "Prime News",
    "Liked Access",
    "Daytime",
    "Prime",
]
# Category tints are all derived from the brand palette (blended toward
# white) so the color system reads as GPS Impact's own, not an arbitrary
# rainbow -- news tiers get the brand's Pale Blue, the liked-show exception
# leans on the brand Red accent, and excluded/informational rows stay a
# neutral Navy-tinted gray.
CATEGORY_FILL = {
    "Early News": PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type="solid"),
    "Noon News": PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type="solid"),
    "Evening News": PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type="solid"),
    "Late News": PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type="solid"),
    "Prime News": PatternFill(start_color="BBCAD8", end_color="BBCAD8", fill_type="solid"),
    "Liked Access": PatternFill(start_color="F6D6D2", end_color="F6D6D2", fill_type="solid"),
    "Daytime": PatternFill(start_color="DADBDF", end_color="DADBDF", fill_type="solid"),
    "Prime": PatternFill(start_color="EEEFF1", end_color="EEEFF1", fill_type="solid"),
}
DAY_MARK_FILL = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="E0E1E4", end_color="E0E1E4", fill_type="solid")
SUBTOTAL_BORDER = Border(top=Side(style="medium", color=NAVY), bottom=Side(style="thin", color=NAVY))
DAY_COLS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DAY_HEADER_LABELS = ["M", "T", "W", "Th", "F", "Sa", "Su"]
INT_FORMAT = "0"  # GRPs and CPP display with no decimals; underlying formula keeps full precision
PCT_FORMAT = "0.0%"  # % Mkt GRPs -- stored as a fraction, Excel's own percent format renders the %

# Fixed column layout for the sample buy grid.
COL_CATEGORY = 1
COL_STATION = 2
COL_PROGRAM = 3
COL_TIME = 4
COL_DAY_FIRST = 5
COL_DAY_LAST = COL_DAY_FIRST + len(DAY_COLS) - 1  # 11 (Sun)
COL_SPOTS = COL_DAY_LAST + 1  # 12
COL_RATE = COL_SPOTS + 1  # 13
COL_RATING = COL_RATE + 1  # 14
COL_CPP = COL_RATING + 1  # 15
COL_WKLY_COST = COL_CPP + 1  # 16
COL_WKLY_GRPS = COL_WKLY_COST + 1  # 17
COL_PCT_MKT = COL_WKLY_GRPS + 1  # 18
NCOLS = COL_PCT_MKT


def _cl(col):
    return get_column_letter(col)


def _add_brand_header(ws, ncols, title_text, subtitle_text=None):
    """GPS Impact banner: logo on its own row, title below it, both on a
    Navy field spanning the sheet's full width, optionally followed by a
    subtitle line still inside the banner. Returns the first free row
    after the banner and its blank spacer."""
    logo_row, title_row = 1, 2
    banner_rows = [logo_row, title_row]

    if LOGO_PATH.exists():
        img = XLImage(str(LOGO_PATH))
        aspect = img.height / img.width
        img.width = 170
        img.height = round(170 * aspect)
        ws.add_image(img, "A1")
    ws.row_dimensions[logo_row].height = 34

    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ncols)
    ws.cell(row=title_row, column=1, value=title_text).font = TITLE_FONT
    ws.row_dimensions[title_row].height = 24

    next_row = title_row + 1
    if subtitle_text is not None:
        subtitle_row = next_row
        banner_rows.append(subtitle_row)
        ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=ncols)
        cell = ws.cell(row=subtitle_row, column=1, value=subtitle_text)
        cell.font = Font(name=FONT_BODY, italic=True, size=10, color=WHITE)
        ws.row_dimensions[subtitle_row].height = 16
        next_row += 1

    for row in banner_rows:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

    return next_row + 1  # skip one blank spacer row after the banner


def _min_to_clock(m):
    h, mm = divmod(int(m) % (24 * 60), 60)
    suffix = "a" if h < 12 else "p"
    h12 = h % 12
    if h12 == 0:
        h12 = 12
    return f"{h12}:{mm:02d}{suffix}"


def _style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _group_avails_into_rows(eligible_avails, spots):
    """One sample buy row per unique (category, station, program, time, rate,
    rating), covering every eligible avail -- not just the ones actually
    bought. The rate/rating are part of the identity because some rate
    cards quote the same program/time slot at different prices across the
    quarter (a rate escalation, or day-pattern variants that happen to
    share identical clock times) -- folding those into a single row would
    let the row's one displayed rate silently mismatch what specific days
    were actually bought at. Days the program doesn't air are omitted;
    days it airs but wasn't bought get an explicit 0 so the buy can be
    edited in place."""
    identities = {}
    for a in eligible_avails:
        key = (a._category, a.station, a.program_name, a.start_min, a.end_min, a.rate, a.rating)
        info = identities.setdefault(key, {"days": set(), "daypart_name": a.daypart_name})
        info["days"].update(a.days)

    bought_counts = defaultdict(lambda: defaultdict(int))
    for s in spots:
        key = (s["category"], s["station"], s["program_name"], s["start_min"], s["end_min"], s["rate"], s["rating"])
        bought_counts[key][s["day"]] += 1

    rows = []
    for (category, station, program, start_min, end_min, rate, rating), info in identities.items():
        counts = bought_counts.get((category, station, program, start_min, end_min, rate, rating), {})
        day_counts = {d: counts.get(d, 0) for d in info["days"]}
        rows.append(
            {
                "category": category,  # internal buy tier -- drives the fill color
                "daypart": daypart_code(info["daypart_name"]),  # normalized to the 2-letter code
                "station": station,
                "program": program,
                "start_min": start_min,
                "end_min": end_min,
                "day_counts": day_counts,
                "rate": rate,
                "rating": rating,
            }
        )
    rows.sort(key=lambda r: (r["station"], r["start_min"], r["program"]))
    return rows


def _write_station_summary_table(ws, result, start_row, stations, station_subtotal_row, grid_total_row):
    """Compact 'totals by station, at a glance' block above the main grid.
    References each station's own subtotal row in the grid below (rather
    than re-deriving the sum), so it stays correct if the grid is edited."""
    spots_col = _cl(COL_SPOTS)
    cost_col = _cl(COL_WKLY_COST)
    grps_col = _cl(COL_WKLY_GRPS)

    ws.cell(row=start_row, column=1, value="Totals by Station (live -- recalculates if you edit the grid below)").font = SECTION_FONT
    header_row = start_row + 1
    headers = ["Station", "Spots/Wk", "Wkly $", "Wkly GRPs", "CPP", "% Mkt GRPs"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    _style_header(ws, header_row, len(headers))

    r = header_row + 1
    for station in stations:
        sub_row = station_subtotal_row[station]
        ws.cell(row=r, column=1, value=station)
        ws.cell(row=r, column=2, value=f"={spots_col}{sub_row}")
        ws.cell(row=r, column=3, value=f"={cost_col}{sub_row}")
        ws.cell(row=r, column=4, value=f"={grps_col}{sub_row}").number_format = INT_FORMAT
        ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)").number_format = INT_FORMAT
        ws.cell(
            row=r, column=6,
            value=f"=IFERROR(D{r}/{grps_col}${grid_total_row},0)",
        ).number_format = PCT_FORMAT
        r += 1

    market_row = r
    ws.cell(row=market_row, column=1, value="MARKET").font = BOLD_FONT
    for col, src_col in ((2, _cl(COL_SPOTS)), (3, _cl(COL_WKLY_COST)), (4, _cl(COL_WKLY_GRPS))):
        cell = ws.cell(row=market_row, column=col, value=f"={src_col}{grid_total_row}")
        cell.font = BOLD_FONT
        if col == 4:
            cell.number_format = INT_FORMAT
    cpp_cell = ws.cell(row=market_row, column=5, value=f"=IFERROR(C{market_row}/D{market_row},0)")
    cpp_cell.font = BOLD_FONT
    cpp_cell.number_format = INT_FORMAT
    market_pct_cell = ws.cell(row=market_row, column=6, value=f"=IFERROR(D{market_row}/{grps_col}${grid_total_row},0)")
    market_pct_cell.font = BOLD_FONT
    market_pct_cell.number_format = PCT_FORMAT

    return market_row + 2  # next free row, with a blank row of padding


def _write_sample_buy_sheet(ws, result, target_demo_label):
    ws.sheet_view.showGridLines = False

    blended_cpp = result.total_cost / result.achieved_grps if result.achieved_grps else 0
    subtitle = (
        f"Target demo: {target_demo_label}   |   "
        f"Weekly GRPs: {result.achieved_grps:.0f} of {result.target_grps:.0f} goal   |   "
        f"Weekly cost: ${result.total_cost:,.0f}   |   Blended CPP: ${blended_cpp:,.0f}   |   "
        f"Edit the day columns below to change the buy -- totals recalculate automatically."
    )
    banner_next_row = _add_brand_header(ws, NCOLS, "Sample Weekly Buy", subtitle)

    rows = _group_avails_into_rows(result.eligible_avails, result.spots)
    stations = sorted({rd["station"] for rd in rows})

    # Layout is fully determined up front so the station-summary formulas
    # (written first, but referencing the grid below) point at the right
    # rows. Each station's block of data rows is followed by a subtotal
    # row and a blank separator row before the next station starts.
    summary_start = banner_next_row
    summary_header_row = summary_start + 1
    summary_market_row = summary_header_row + 1 + len(stations)
    header_row = summary_market_row + 2

    station_data_range = {}  # station -> (first_row, last_row)
    station_subtotal_row = {}
    r = header_row + 1
    for station in stations:
        n = sum(1 for rd in rows if rd["station"] == station)
        station_data_range[station] = (r, r + n - 1)
        r += n
        station_subtotal_row[station] = r
        r += 1  # subtotal row
        r += 1  # blank separator row
    total_row = r

    _write_station_summary_table(
        ws, result, summary_start, stations, station_subtotal_row, total_row
    )

    headers = (
        ["Daypart", "Station", "Program", "Time"]
        + DAY_HEADER_LABELS
        + ["Spots/Wk", "Rate", "Rating", "CPP", "Wkly $", "Wkly GRPs", "% Mkt GRPs"]
    )
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    _style_header(ws, header_row, len(headers))

    day_first_letter = _cl(COL_DAY_FIRST)
    day_last_letter = _cl(COL_DAY_LAST)
    rate_letter = _cl(COL_RATE)
    rating_letter = _cl(COL_RATING)
    spots_letter = _cl(COL_SPOTS)
    grps_letter = _cl(COL_WKLY_GRPS)

    r = header_row + 1
    for row_data in rows:
        station = row_data["station"]
        fill = CATEGORY_FILL.get(row_data["category"])
        for col, value in (
            (COL_CATEGORY, row_data["daypart"]),
            (COL_STATION, row_data["station"]),
            (COL_PROGRAM, row_data["program"]),
        ):
            cell = ws.cell(row=r, column=col, value=value)
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER

        time_label = f"{_min_to_clock(row_data['start_min'])}-{_min_to_clock(row_data['end_min'])}"
        cell = ws.cell(row=r, column=COL_TIME, value=time_label)
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        for i, day in enumerate(DAY_COLS):
            col = COL_DAY_FIRST + i
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if day in row_data["day_counts"]:
                count = row_data["day_counts"][day]
                cell.value = count
                if count > 0:
                    cell.fill = DAY_MARK_FILL
                    cell.font = DAY_BOUGHT_FONT
                else:
                    if fill:
                        cell.fill = fill
                    cell.font = DAY_UNBOUGHT_FONT
            elif fill:
                cell.fill = fill

        cell = ws.cell(row=r, column=COL_SPOTS, value=f"=SUM({day_first_letter}{r}:{day_last_letter}{r})")
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        cell = ws.cell(row=r, column=COL_RATE, value=row_data["rate"])
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        cell = ws.cell(row=r, column=COL_RATING, value=row_data["rating"])
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        cell = ws.cell(row=r, column=COL_CPP, value=f"=IFERROR({rate_letter}{r}/{rating_letter}{r},0)")
        cell.number_format = INT_FORMAT
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        cell = ws.cell(row=r, column=COL_WKLY_COST, value=f"={spots_letter}{r}*{rate_letter}{r}")
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        cell = ws.cell(row=r, column=COL_WKLY_GRPS, value=f"={spots_letter}{r}*{rating_letter}{r}")
        cell.number_format = INT_FORMAT
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        cell = ws.cell(
            row=r, column=COL_PCT_MKT,
            value=f"=IFERROR({grps_letter}{r}/{grps_letter}${total_row},0)",
        )
        cell.number_format = PCT_FORMAT
        if fill:
            cell.fill = fill
        cell.border = THIN_BORDER

        is_last_row_for_station = r == station_data_range[station][1]
        r += 1

        if is_last_row_for_station:
            sub_row = station_subtotal_row[station]
            data_first, data_last = station_data_range[station]
            ws.cell(row=sub_row, column=COL_CATEGORY, value=f"{station} TOTAL")
            ws.cell(row=sub_row, column=COL_STATION, value=station)
            ws.cell(
                row=sub_row, column=COL_SPOTS,
                value=f"=SUM({spots_letter}{data_first}:{spots_letter}{data_last})",
            )
            ws.cell(
                row=sub_row, column=COL_CPP,
                value=f"=IFERROR({_cl(COL_WKLY_COST)}{sub_row}/{grps_letter}{sub_row},0)",
            ).number_format = INT_FORMAT
            ws.cell(
                row=sub_row, column=COL_WKLY_COST,
                value=f"=SUM({_cl(COL_WKLY_COST)}{data_first}:{_cl(COL_WKLY_COST)}{data_last})",
            )
            ws.cell(
                row=sub_row, column=COL_WKLY_GRPS,
                value=f"=SUM({grps_letter}{data_first}:{grps_letter}{data_last})",
            ).number_format = INT_FORMAT
            ws.cell(
                row=sub_row, column=COL_PCT_MKT,
                value=f"=IFERROR({grps_letter}{sub_row}/{grps_letter}${total_row},0)",
            ).number_format = PCT_FORMAT
            for col in range(1, NCOLS + 1):
                cell = ws.cell(row=sub_row, column=col)
                cell.font = BOLD_FONT
                cell.fill = SUBTOTAL_FILL
                cell.border = SUBTOTAL_BORDER
            r = sub_row + 2  # skip the subtotal row and its blank separator

    subtotal_refs = lambda col: ",".join(f"{col}{station_subtotal_row[s]}" for s in stations)
    ws.cell(row=total_row, column=COL_CATEGORY, value="MARKET TOTAL").font = BOLD_FONT
    ws.cell(
        row=total_row, column=COL_SPOTS,
        value=f"=SUM({subtotal_refs(spots_letter)})",
    ).font = BOLD_FONT
    ws.cell(
        row=total_row, column=COL_WKLY_COST,
        value=f"=SUM({subtotal_refs(_cl(COL_WKLY_COST))})",
    ).font = BOLD_FONT
    total_grps_cell = ws.cell(
        row=total_row, column=COL_WKLY_GRPS,
        value=f"=SUM({subtotal_refs(grps_letter)})",
    )
    total_grps_cell.font = BOLD_FONT
    total_grps_cell.number_format = INT_FORMAT
    total_cpp_cell = ws.cell(
        row=total_row, column=COL_CPP,
        value=f"=IFERROR({_cl(COL_WKLY_COST)}{total_row}/{grps_letter}{total_row},0)",
    )
    total_cpp_cell.font = BOLD_FONT
    total_cpp_cell.number_format = INT_FORMAT
    total_pct_cell = ws.cell(
        row=total_row, column=COL_PCT_MKT,
        value=f"=IFERROR({grps_letter}{total_row}/{grps_letter}{total_row},0)",
    )
    total_pct_cell.font = BOLD_FONT
    total_pct_cell.number_format = PCT_FORMAT

    ws.freeze_panes = ws.cell(row=header_row + 1, column=COL_DAY_FIRST)

    widths = [13, 8, 36, 15] + [4] * len(DAY_COLS) + [9, 8, 8, 8, 11, 11, 11]
    _autofit(ws, widths)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)
    ws.print_title_rows = f"{header_row}:{header_row}"


def write_workbook(result, path, target_demo_label="Adults 35+"):
    wb = Workbook()

    # ---- Sample Buy: the single-page, at-a-glance, editable view ----
    ws = wb.active
    ws.title = "Sample Buy"
    _write_sample_buy_sheet(ws, result, target_demo_label)

    # ---- Market Summary ----
    ws = wb.create_sheet("Market Summary")
    SUMMARY_NCOLS = 5
    banner_next_row = _add_brand_header(ws, SUMMARY_NCOLS, "Sample Weekly Buy -- Market Summary")

    row = banner_next_row
    ws.cell(row=row, column=1, value=f"Target demo: {target_demo_label}")
    row += 1
    ws.cell(row=row, column=1, value=f"Target weekly GRPs: {result.target_grps:.0f}")
    row += 1
    ws.cell(row=row, column=1, value=f"Achieved weekly GRPs: {result.achieved_grps:.0f}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total weekly cost: ${result.total_cost:,.0f}")
    row += 1
    blended_cpp = result.total_cost / result.achieved_grps if result.achieved_grps else 0
    ws.cell(row=row, column=1, value=f"Blended market CPP: ${blended_cpp:,.0f}")
    row += 1
    ws.cell(row=row, column=1, value=f"Total spots/week: {len(result.spots)}")
    row += 2
    if result.warnings:
        ws.cell(row=row, column=1, value="Notes").font = SECTION_FONT
        row += 1
        for w in result.warnings:
            ws.cell(row=row, column=1, value=f"- {w}")
            row += 1
        row += 1

    if result.outlier_avails:
        ws.cell(row=row, column=1, value="News avails deprioritized as CPP outliers").font = SECTION_FONT
        row += 1
        headers = ["Station", "Program", "CPP"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1
        for station, program, cpp in result.outlier_avails:
            ws.cell(row=row, column=1, value=station)
            ws.cell(row=row, column=2, value=program)
            ws.cell(row=row, column=3, value=cpp)
            row += 1
        row += 1

    # breakdown by category
    ws.cell(row=row, column=1, value="Breakdown by Daypart Category").font = SECTION_FONT
    row += 1
    headers = ["Category", "Spots", "Weekly GRPs", "Weekly Cost", "CPP"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, len(headers))
    row += 1
    by_cat = defaultdict(lambda: {"spots": 0, "grps": 0.0, "cost": 0.0})
    for s in result.spots:
        b = by_cat[s["category"]]
        b["spots"] += 1
        b["grps"] += s["rating"]
        b["cost"] += s["rate"]
    for cat in CATEGORY_ORDER:
        if cat not in by_cat:
            continue
        b = by_cat[cat]
        cpp = b["cost"] / b["grps"] if b["grps"] else 0
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=b["spots"])
        ws.cell(row=row, column=3, value=round(b["grps"]))
        ws.cell(row=row, column=4, value=round(b["cost"], 0))
        ws.cell(row=row, column=5, value=round(cpp))
        row += 1
    row += 1

    # breakdown by station
    ws.cell(row=row, column=1, value="Breakdown by Station").font = SECTION_FONT
    row += 1
    headers = ["Station", "Spots", "Weekly GRPs", "Weekly Cost", "CPP"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, len(headers))
    row += 1
    by_station = defaultdict(lambda: {"spots": 0, "grps": 0.0, "cost": 0.0})
    for s in result.spots:
        b = by_station[s["station"]]
        b["spots"] += 1
        b["grps"] += s["rating"]
        b["cost"] += s["rate"]
    for station in sorted(by_station):
        b = by_station[station]
        cpp = b["cost"] / b["grps"] if b["grps"] else 0
        ws.cell(row=row, column=1, value=station)
        ws.cell(row=row, column=2, value=b["spots"])
        ws.cell(row=row, column=3, value=round(b["grps"]))
        ws.cell(row=row, column=4, value=round(b["cost"], 0))
        ws.cell(row=row, column=5, value=round(cpp))
        row += 1

    _autofit(ws, [42, 12, 14, 14, 10])

    wb.save(path)
